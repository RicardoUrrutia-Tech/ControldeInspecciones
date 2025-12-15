import io
import re
from pathlib import Path

import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt


# =========================================================
# Configuración general (solo 1 vez y al inicio)
# =========================================================
st.set_page_config(
    layout="wide"
)

# =========================================================
# SSO / Acceso restringido (Google OAuth vía Streamlit)
# =========================================================
CORP_DOMAIN = "@cabify.com"

def require_login_and_domain():
    if "auth" not in st.secrets:
        st.error("No se encontró configuración [auth] en Secrets de Streamlit Cloud.")
        st.stop()

    if not getattr(st.user, "is_logged_in", False):
        st.title("🔐 Acceso restringido")
        st.write("Debes iniciar sesión con tu cuenta corporativa para usar esta aplicación.")
        st.button("Iniciar sesión con Google", on_click=st.login)
        st.stop()

    email = (getattr(st.user, "email", "") or "").strip().lower()
    if not email.endswith(CORP_DOMAIN):
        st.title("🔐 Acceso restringido")
        st.error(f"Debes ingresar con una cuenta corporativa ({CORP_DOMAIN}).")
        st.write(f"Sesión detectada: {email or '(sin email)'}")
        st.button("Cerrar sesión", on_click=st.logout)
        st.stop()

require_login_and_domain()

# =========================================================
# Estilo Cabify (CSS)
# =========================================================
st.markdown("""
<style>
/* Fondo general */
.stApp { background-color: #FAF8FE; }

/* Títulos */
h1, h2, h3 { color: #1F123F; }
h4, h5, h6 { color: #4A2B8D; }

/* Sidebar */
section[data-testid="stSidebar"] { background-color: #1F123F; }
section[data-testid="stSidebar"] * { color: #FAF8FE !important; }

/* Botones informales (st.button / download) */
.stButton>button, .stDownloadButton>button {
    background-color: #7145D6;
    color: #FFFFFF;
    border-radius: 10px;
    border: none;
    padding: 0.55rem 0.9rem;
    font-weight: 700;
}
.stButton>button:hover, .stDownloadButton>button:hover {
    background-color: #5B34AC;
    color: #FFFFFF;
}

/* Link buttons */
a[data-testid="stLinkButton"] {
    background-color: #8A6EE4 !important;
    color: #FFFFFF !important;
    border-radius: 10px !important;
    padding: 0.55rem 0.9rem !important;
    font-weight: 800 !important;
    border: none !important;
}

/* Alerts */
div[data-testid="stAlert"] { border-radius: 10px; }

/* Separadores */
hr { border-top: 1px solid #DFDAF8; }
</style>
""", unsafe_allow_html=True)

with st.sidebar:
    st.write("### Sesión")
    st.write(f"Conectado como: **{st.user.email}**")
    st.button("Cerrar sesión", on_click=st.logout)

# =========================================================
# Portada (imagen adjunta en el repo)
# =========================================================
# Guarda tu imagen en: assets/portada_aeropuerto.png  (o cambia el nombre aquí)
PORTADA_PATH = Path("assets/portada_aeropuerto.png")
if PORTADA_PATH.exists():
    st.image(str(PORTADA_PATH), use_container_width=True)

# =========================================================
# Links de descarga (Drive)
# =========================================================
PATENTES_EXPORT_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1DSZA9DJkxWHBIOMyulW3638TF46Udq7tCWh489UAOAs/"
    "export?format=xlsx&gid=1266707607"
)

INSPECCIONES_EXPORT_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1UR2GTh6l4nwmx4DdY6yX_BS7EF1l-LYHtzHQdkYx03c/"
    "export?format=xlsx"
)

# =========================================================
# Utilidades
# =========================================================
def normalize_headers(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = (
        df.columns.astype(str)
        .str.replace("\u00a0", " ", regex=False)   # NBSP
        .str.replace(r"\s+", " ", regex=True)     # colapsa espacios/saltos de línea
        .str.strip()
    )
    return df

def normalize_plate(x: str) -> str:
    if pd.isna(x):
        return ""
    s = str(x).strip().upper()
    s = re.sub(r"[^A-Z0-9]", "", s)
    return s

def try_parse_date(series: pd.Series) -> pd.Series:
    dt = pd.to_datetime(series, errors="coerce", dayfirst=True)
    mask = dt.isna()
    if mask.any():
        numeric = pd.to_numeric(series[mask], errors="coerce")
        dt2 = pd.to_datetime(numeric, unit="D", origin="1899-12-30", errors="coerce")
        dt.loc[mask] = dt2
    return dt

def traffic_light(days, green_max, yellow_max):
    if pd.isna(days):
        return "⚫ Sin inspección"
    if days <= green_max:
        return "🟢 OK"
    if days <= yellow_max:
        return "🟡 Alerta"
    return "🔴 Crítico"

def compliance_label(x):
    """
    Devuelve:
    - 'Cumple' si valor == 100
    - 'No Cumple' si valor < 100
    - NaN si no hay valor
    """
    if pd.isna(x):
        return pd.NA
    v = pd.to_numeric(str(x).strip().replace("%", ""), errors="coerce")
    if pd.isna(v):
        return pd.NA
    return "Cumple" if v >= 100 else "No Cumple"

def style_semaforo(val: str) -> str:
    if val == "🟢 OK":
        return "color: #0C936B; font-weight: 800"
    if val == "🟡 Alerta":
        return "color: #EFBD03; font-weight: 800"
    if val == "🔴 Crítico":
        return "color: #E74A41; font-weight: 800"
    if val == "⚫ Sin inspección":
        return "color: #362065; font-weight: 800"
    return ""

def style_cumplimiento(val: str) -> str:
    if val == "Cumple":
        return "color: #0C936B; font-weight: 800"
    if val == "No Cumple":
        return "color: #E74A41; font-weight: 800"
    return ""

def to_excel_bytes(df: pd.DataFrame) -> bytes:
    """
    Export Cabify:
    - Header morado
    - Congelar primera fila
    - Estilo básico de Semáforo + Cumplimientos (font)
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="resultado")
        ws = writer.sheets["resultado"]

        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="7145D6")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        ws.freeze_panes = "A2"

        # Ajuste simple de ancho
        for i, col_name in enumerate(df.columns, start=1):
            width = min(max(len(str(col_name)) + 2, 14), 40)
            ws.column_dimensions[get_column_letter(i)].width = width

        # Colores
        font_ok = Font(color="0C936B", bold=True)
        font_alert = Font(color="EFBD03", bold=True)
        font_crit = Font(color="E74A41", bold=True)
        font_neutral = Font(color="362065", bold=True)

        idx = {name: j + 1 for j, name in enumerate(df.columns)}
        sem_col = idx.get("Semáforo")
        ce_col = idx.get("Cumplimiento Exterior")
        ci_col = idx.get("Cumplimiento Interior")
        cc_col = idx.get("Cumplimiento Conductor")

        def set_font(r, c, font):
            if c is None:
                return
            ws.cell(row=r, column=c).font = font

        for r in range(2, ws.max_row + 1):
            # Semáforo
            if sem_col is not None:
                v = ws.cell(row=r, column=sem_col).value
                if v == "🟢 OK":
                    set_font(r, sem_col, font_ok)
                elif v == "🟡 Alerta":
                    set_font(r, sem_col, font_alert)
                elif v == "🔴 Crítico":
                    set_font(r, sem_col, font_crit)
                elif v == "⚫ Sin inspección":
                    set_font(r, sem_col, font_neutral)

            # Cumplimientos
            for ccol in [ce_col, ci_col, cc_col]:
                if ccol is None:
                    continue
                v = ws.cell(row=r, column=ccol).value
                if v == "Cumple":
                    set_font(r, ccol, font_ok)
                elif v == "No Cumple":
                    set_font(r, ccol, font_crit)

    return output.getvalue()

# =========================================================
# UI
# =========================================================
st.title("🚗 Última inspección por patente – Aeropuerto")
st.caption(
    "Esta app mantiene los Google Sheets **privados**. "
    "La descarga se hace en tu navegador usando tu **sesión corporativa**, y luego subes los Excel aquí."
)

st.divider()

# Paso 1: Descargas
st.subheader("Paso 1) Descarga los Excel")
st.info("Abre cada botón y descarga el Excel. Luego vuelve a esta app para subir ambos archivos.")

c1, c2 = st.columns(2)
with c1:
    st.link_button("📄 Descargar Base Patentes (Excel)", PATENTES_EXPORT_URL)
with c2:
    st.link_button("📝 Descargar Inspecciones (Excel)", INSPECCIONES_EXPORT_URL)

st.divider()

# Paso 2: Upload
st.subheader("Paso 2) Sube los dos archivos descargados")

col_u1, col_u2 = st.columns(2)
with col_u1:
    f_pat = st.file_uploader("Sube **Base Patentes Aeropuerto** (Excel)", type=["xlsx", "xls"])
with col_u2:
    f_insp = st.file_uploader("Sube **Inspecciones Aeropuerto** (Excel)", type=["xlsx", "xls"])

if not f_pat or not f_insp:
    st.info("👆 Sube ambos archivos para continuar al Paso 3.")
    st.stop()

df_pat = normalize_headers(pd.read_excel(f_pat))
df_insp = normalize_headers(pd.read_excel(f_insp))

# Validaciones mínimas
if "REG PLATE" not in df_pat.columns:
    st.error(
        "❌ El archivo de Patentes no parece correcto: falta la columna **REG PLATE**.\n"
        "Descárgalo desde el botón y vuelve a subirlo."
    )
    with st.expander("Ver columnas detectadas (Patentes)"):
        st.write(list(df_pat.columns))
    st.stop()

required_insp = [
    "Fecha",
    "Patente del Vehículo",
    "Cumplimiento Exterior",
    "Cumplimiento Interior",
    "Cumplimiento Conductor",
]
missing = [c for c in required_insp if c not in df_insp.columns]
if missing:
    st.error(
        "❌ El archivo de Inspecciones no parece correcto: faltan columnas clave:\n- "
        + "\n- ".join(missing)
        + "\n\nDescárgalo desde el botón y vuelve a subirlo."
    )
    with st.expander("Ver columnas detectadas (Inspecciones)"):
        st.write(list(df_insp.columns))
    st.stop()

st.success("✅ Archivos cargados correctamente.")
st.caption("Si subes un Excel y la app dice que faltan columnas, revisa que hayas descargado el archivo correcto desde los botones.")
st.divider()

# Paso 3: Resultados
st.subheader("Paso 3) Resultados")

with st.expander("⚙️ Configuración semáforo (días desde la última inspección)", expanded=True):
    a1, a2 = st.columns(2)
    with a1:
        green_max = st.number_input("🟢 OK hasta (días)", min_value=0, value=7, step=1)
    with a2:
        yellow_max = st.number_input("🟡 Alerta hasta (días)", min_value=0, value=30, step=1)

    if yellow_max < green_max:
        st.warning("El umbral 🟡 (Alerta) no puede ser menor que 🟢 (OK). Se ajustó automáticamente.")
        yellow_max = green_max

# Procesamiento
df_pat = df_pat.copy()
df_insp = df_insp.copy()

df_pat["plate_norm"] = df_pat["REG PLATE"].apply(normalize_plate)
df_insp["plate_norm"] = df_insp["Patente del Vehículo"].apply(normalize_plate)
df_insp["Fecha_dt"] = try_parse_date(df_insp["Fecha"])

# Última inspección por patente
df_last = (
    df_insp[df_insp["plate_norm"] != ""]
    .loc[df_insp["Fecha_dt"].notna()]
    .sort_values(["plate_norm", "Fecha_dt"], ascending=[True, False])
    .groupby("plate_norm", as_index=False)
    .first()
)

# Cumplimiento -> Cumple / No Cumple
for c in ["Cumplimiento Exterior", "Cumplimiento Interior", "Cumplimiento Conductor"]:
    df_last[c] = df_last[c].apply(compliance_label)

df_last = df_last[
    [
        "plate_norm",
        "Fecha_dt",
        "Cumplimiento Exterior",
        "Cumplimiento Interior",
        "Cumplimiento Conductor",
    ]
].rename(columns={"Fecha_dt": "Última Fecha Inspección"})

df = df_pat.merge(df_last, on="plate_norm", how="left")

today = pd.Timestamp.today().normalize()
df["Inspeccionado"] = df["Última Fecha Inspección"].notna()
df["Días desde última inspección"] = (today - pd.to_datetime(df["Última Fecha Inspección"])).dt.days
df["Semáforo"] = df["Días desde última inspección"].apply(lambda x: traffic_light(x, green_max, yellow_max))

# KPIs
total_pat = int(df["REG PLATE"].notna().sum())
inspected = int(df["Inspeccionado"].sum())
never = total_pat - inspected

k1, k2, k3, k4 = st.columns(4)
k1.metric("Total Patentes", f"{total_pat:,}".replace(",", "."))
k2.metric("Con inspección", f"{inspected:,}".replace(",", "."))
k3.metric("Sin inspección", f"{never:,}".replace(",", "."))
k4.metric("Fecha de cálculo", today.strftime("%Y-%m-%d"))

st.divider()

# Top tablas
left, right = st.columns(2)

with left:
    st.subheader("🧾 Top sin inspección (primeras 20)")
    base_cols = [c for c in ["REG PLATE", "Flota", "Company", "Marca", "Modelo", "Color"] if c in df.columns]
    cols_never = base_cols + ["Semáforo"]
    st.dataframe(
        df[df["Inspeccionado"] == False][cols_never].head(20),
        use_container_width=True,
        hide_index=True
    )

with right:
    st.subheader("🕰️ Top inspecciones más antiguas (primeras 20)")
    cols_old = [
        "REG PLATE",
        "Última Fecha Inspección",
        "Días desde última inspección",
        "Semáforo",
        "Cumplimiento Exterior",
        "Cumplimiento Interior",
        "Cumplimiento Conductor",
    ]
    cols_old = [c for c in cols_old if c in df.columns]
    df_old = (
        df[df["Inspeccionado"] == True]
        .sort_values("Días desde última inspección", ascending=False)[cols_old]
        .head(20)
    )
    st.dataframe(
        df_old.style
            .applymap(style_semaforo, subset=["Semáforo"])
            .applymap(style_cumplimiento, subset=[c for c in ["Cumplimiento Exterior", "Cumplimiento Interior", "Cumplimiento Conductor"] if c in df_old.columns]),
        use_container_width=True,
        hide_index=True
    )

st.divider()

# Gráfico distribución (compacto + barras moradul)
st.subheader("📊 Distribución de días desde última inspección (solo inspeccionados)")

vals = df.loc[df["Inspeccionado"] == True, "Días desde última inspección"].dropna()
if vals.empty:
    st.info("No hay inspecciones con fecha válida para graficar.")
else:
    bins = st.slider("Número de bins (barras)", min_value=5, max_value=60, value=20, step=1)
    g_left, g_right = st.columns([2, 3])  # compacto
    with g_left:
        fig = plt.figure(figsize=(6, 3))
        plt.hist(vals, bins=bins, color="#4A2B8D")  # ← moradul Cabify
        plt.xlabel("Días")
        plt.ylabel("Cantidad")
        st.pyplot(fig, use_container_width=True)

st.divider()

# Tabla principal filtrable
st.subheader("📋 Resultado completo (filtrable)")

f1, f2, f3, f4 = st.columns([1, 1, 1, 2])
with f1:
    only_never = st.checkbox("Solo SIN inspección", value=False)
with f2:
    only_inspected = st.checkbox("Solo CON inspección", value=False)
with f3:
    sem_filter = st.selectbox("Filtrar por semáforo", ["(Todos)", "⚫ Sin inspección", "🔴 Crítico", "🟡 Alerta", "🟢 OK"])
with f4:
    q = st.text_input("Buscar patente (ej: ABCD12)", value="").strip().upper()

df_show = df.copy()

if only_never and not only_inspected:
    df_show = df_show[df_show["Inspeccionado"] == False]
elif only_inspected and not only_never:
    df_show = df_show[df_show["Inspeccionado"] == True]

if sem_filter != "(Todos)":
    df_show = df_show[df_show["Semáforo"] == sem_filter]

if q:
    qn = normalize_plate(q)
    df_show = df_show[df_show["plate_norm"].str.contains(qn, na=False)]

cols_view = [
    "REG PLATE",
    "Semáforo",
    "Última Fecha Inspección",
    "Días desde última inspección",
    "Cumplimiento Exterior",
    "Cumplimiento Interior",
    "Cumplimiento Conductor",
    "Inspeccionado",
]
optional_base_cols = ["Marca", "Modelo", "Color", "Flota", "Company"]
for c in optional_base_cols:
    if c in df_show.columns and c not in cols_view:
        cols_view.insert(1, c)

cols_view = [c for c in cols_view if c in df_show.columns]

styler = (
    df_show[cols_view].style
    .applymap(style_semaforo, subset=["Semáforo"] if "Semáforo" in cols_view else None)
    .applymap(
        style_cumplimiento,
        subset=[c for c in ["Cumplimiento Exterior", "Cumplimiento Interior", "Cumplimiento Conductor"] if c in cols_view]
    )
)

st.dataframe(styler, use_container_width=True, hide_index=True)

st.download_button(
    "⬇️ Descargar resultado filtrado (Excel)",
    data=to_excel_bytes(df_show[cols_view]),
    file_name="ultima_inspeccion_por_patente.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

